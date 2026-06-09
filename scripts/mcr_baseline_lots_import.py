#!/usr/bin/env python3
"""
MCR Baseline Lots Importer (one-shot, idempotent).

Re-parses MCR's April 2026 inventory spreadsheet (GREEN COFFEE sheet) to create
proper per-lot `coffee_inventory_purchased` rows with entry_method='baseline'.

Replaces 14-21 generic per-origin synthetic baselines from Phase 1.

Usage:
    python3 scripts/mcr_baseline_lots_import.py [--dry-run]

The script:
  1. Parses every valid lot row in GREEN COFFEE
  2. Matches each row to existing coffee_source (creates if missing)
  3. Matches each row to existing coffee_inventory origin
  4. Deletes the existing entry_method='baseline' rows for MCR
  5. Inserts one per-lot baseline row per spreadsheet lot
  6. Calls recalculate_origin_total_stock() for every touched origin
  7. Prints an audit summary
"""
from __future__ import annotations

import argparse
import os
import re
import sys
import uuid
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import Optional

import openpyxl
import psycopg2
import psycopg2.extras

# --- Configuration ----------------------------------------------------------

COMPANY_ID = "9ShiyDAXhV"
FACILITY_ID = "5cc581b9-2803-42c2-98de-0ba16ae42f8e"
SPREADSHEET = Path(
    "/Users/wanderingaloha/my-supabase-project/MCR/strata migration/"
    "April 2026 - Inventory Spreadsheet.xlsx"
)
SHEET = "GREEN COFFEE"

PG_DSN = (
    "host=db.pwpslalerytymorcodlv.supabase.co "
    "port=5432 user=postgres dbname=postgres "
    "password=SDH-h3FNHXSrxj-"
)

# Rows where column A matches one of these (case-insensitive, stripped) are
# category headers or totals — skip them.
CATEGORY_HEADERS = {
    "hawaiian",
    "maui",
    "decaf",
    "international",
    "international organic",
    "international decaf",
}


def is_total_row(name: str) -> bool:
    lower = name.lower().strip()
    return lower.endswith("total") or lower.endswith("total:")


# --- Spreadsheet parsing ---------------------------------------------------

@dataclass
class LotRow:
    excel_row: int
    coffee_name: str        # Column A — coffee identity (matches coffee_source.coffee_name)
    lot_ref_raw: object     # Column B — bag #, lot, contract ref
    supplier_ref: str       # Column D — "Royal Coffee #37040" etc
    cost_lb: float          # Column E
    bag_size_lbs: float     # Column F
    bag_count: float        # Column G
    partial_lbs: float      # Column H
    total_lbs: float        # Column I
    category: str           # last seen category header

    @property
    def lot_id(self) -> str:
        """Normalize lot ref to a stable lot_id string."""
        raw = self.lot_ref_raw
        if raw is None or str(raw).strip() == "":
            # No lot ref — fall back to supplier ref or coffee+row
            if self.supplier_ref:
                return f"{self.supplier_ref.strip()}-r{self.excel_row}"
            return f"BASELINE-r{self.excel_row}"
        s = str(raw).strip()
        # strip common label prefixes
        s = re.sub(r"^(Lot#?\s*:?\s*|INV\s+|Contract#?\s*:?\s*|Item:\s*)", "", s, flags=re.IGNORECASE)
        s = s.strip()
        if not s:
            return f"BASELINE-r{self.excel_row}"
        return s


def parse_spreadsheet() -> list[LotRow]:
    wb = openpyxl.load_workbook(SPREADSHEET, data_only=True)
    ws = wb[SHEET]

    lots: list[LotRow] = []
    current_category = ""
    skipped: list[tuple[int, str]] = []

    for r in range(4, ws.max_row + 1):
        a = ws.cell(r, 1).value  # ITEM / coffee name
        b = ws.cell(r, 2).value  # BAG # / lot
        d = ws.cell(r, 4).value  # ITEM DESCRIPTION / supplier
        e = ws.cell(r, 5).value  # $/LB
        f = ws.cell(r, 6).value  # LBS (bag size)
        g = ws.cell(r, 7).value  # BAGS
        h = ws.cell(r, 8).value  # PARTIAL
        i = ws.cell(r, 9).value  # TOTAL LBS

        if a is None:
            continue
        name = str(a).strip()
        if not name:
            continue

        lower = name.lower().rstrip()
        if lower in CATEGORY_HEADERS:
            current_category = name
            continue
        if is_total_row(name):
            continue

        # Need cost + bag size + (bags or total lbs) to be a real lot row
        if e is None or f is None:
            skipped.append((r, f"no cost or bag size: name={name!r}"))
            continue

        try:
            cost_lb = float(e)
            bag_size_lbs = float(f)
        except (TypeError, ValueError):
            skipped.append((r, f"cost/bag-size not numeric: {e!r}/{f!r}"))
            continue

        bag_count = float(g) if g not in (None, "") else 0.0
        partial_lbs = float(h) if h not in (None, "") else 0.0
        total_lbs = float(i) if i not in (None, "") else 0.0

        # Recompute total if missing or off
        computed = bag_count * bag_size_lbs + partial_lbs
        if total_lbs == 0 and computed > 0:
            total_lbs = computed

        # Skip rows that hold zero inventory — no on-hand baseline to record
        if total_lbs <= 0:
            skipped.append((r, f"zero on-hand: {name}"))
            continue

        lots.append(
            LotRow(
                excel_row=r,
                coffee_name=name,
                lot_ref_raw=b,
                supplier_ref=str(d).strip() if d else "",
                cost_lb=cost_lb,
                bag_size_lbs=bag_size_lbs,
                bag_count=bag_count,
                partial_lbs=partial_lbs,
                total_lbs=total_lbs,
                category=current_category,
            )
        )

    return lots, skipped


# --- Matching helpers ------------------------------------------------------

def normalize(s: str) -> str:
    """Mirror lib/artisan/contentMatch.ts normalize: lowercase, strip punct, collapse whitespace."""
    s = s.lower()
    s = re.sub(r"['`’]", "", s)
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


# Spreadsheet category → origin name (case sensitive match to DB)
CATEGORY_TO_ORIGIN_HINT = {
    "Hawaiian": None,  # ambiguous — must derive from coffee name
    "Maui": "Maui",
    "Decaf": None,
    "International": None,
    "International Organic": None,
    "International Decaf": None,
}

# Map first-word/keyword in coffee name → origin display name
ORIGIN_KEYWORD_MAP = [
    ("ka'u", "Ka'u"),
    ("kau", "Ka'u"),
    ("kona", "Kona"),
    ("hawaii", "Kona"),  # "Hawaii No.3 (Kona)"
    ("maui", "Maui"),
    ("brazil", "Brazil"),
    ("colombia", "Colombia"),
    ("colombian", "Colombia"),
    ("costa rica", "Costa Rica"),
    ("el salvador", "El Salvador"),
    ("guatemala", "Guatemala"),
    ("honduras", "Honduras"),
    ("mexico", "Mexico"),
    ("nicaragu", "Nicaragua"),  # "Nicaragu" used in sheet
    ("papa new guinea", "Papua New Guinea"),
    ("papua new guinea", "Papua New Guinea"),
    ("png", "Papua New Guinea"),
    ("peru", "Peru"),
    ("sumatra", "Sumatra"),
    ("timor", "Timor"),
    ("yemen", "Yemen"),
]


def guess_origin_name(coffee_name: str) -> Optional[str]:
    n = normalize(coffee_name)
    for kw, origin in ORIGIN_KEYWORD_MAP:
        if kw in n:
            return origin
    return None


def find_coffee_source(
    cur,
    coffee_name: str,
    origin_id: str,
    coffee_sources: list[dict],
) -> Optional[str]:
    """Return coffee_source_id matching coffee_name under origin_id, or None."""
    target = normalize(coffee_name)
    # 1. Exact normalized match within same origin
    for cs in coffee_sources:
        if cs["origin_id"] == origin_id and normalize(cs["coffee_name"]) == target:
            return cs["coffee_source_id"]
    # 2. Best containment match within same origin
    candidates = [cs for cs in coffee_sources if cs["origin_id"] == origin_id]
    for cs in candidates:
        cs_norm = normalize(cs["coffee_name"])
        if cs_norm and (cs_norm in target or target in cs_norm):
            # Require at least 60% token overlap to accept
            tt = set(target.split())
            ct = set(cs_norm.split())
            if not tt or not ct:
                continue
            overlap = len(tt & ct) / max(len(tt), len(ct))
            if overlap >= 0.6:
                return cs["coffee_source_id"]
    return None


# --- Main pipeline ---------------------------------------------------------

def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--dry-run", action="store_true",
                        help="Parse + match but do not write to the database.")
    args = parser.parse_args()

    print(f"Parsing {SPREADSHEET} sheet {SHEET!r}…")
    lots, skipped = parse_spreadsheet()
    print(f"  Parsed {len(lots)} lot rows, skipped {len(skipped)} rows.")

    print("Connecting to DB…")
    conn = psycopg2.connect(PG_DSN)
    conn.autocommit = False
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    # Load existing MCR origins
    cur.execute(
        "SELECT origin_id, origin FROM coffee_inventory "
        "WHERE company_id = %s AND facility_id = %s",
        (COMPANY_ID, FACILITY_ID),
    )
    origins = cur.fetchall()
    origin_by_name = {o["origin"].strip(): o["origin_id"] for o in origins}
    print(f"  Loaded {len(origins)} MCR origins.")

    # Load existing MCR coffee_sources
    cur.execute(
        "SELECT coffee_source_id, coffee_name, origin_id, bag_size "
        "FROM coffee_source WHERE company_id = %s",
        (COMPANY_ID,),
    )
    coffee_sources = [dict(r) for r in cur.fetchall()]
    print(f"  Loaded {len(coffee_sources)} MCR coffee_sources.")

    # Load known bag sizes
    cur.execute(
        "SELECT bag_size_id FROM bag_sizes "
        "WHERE company_id IS NULL OR company_id = %s",
        (COMPANY_ID,),
    )
    bag_size_ids = {r["bag_size_id"] for r in cur.fetchall()}
    print(f"  Loaded {len(bag_size_ids)} known bag_sizes ids.")

    new_coffee_sources: list[tuple[str, str]] = []  # (coffee_source_id, coffee_name)
    new_bag_sizes: list[str] = []

    # Resolve every lot row: origin_id, coffee_source_id, bag_size text
    resolved: list[dict] = []
    unresolved: list[tuple[int, str]] = []

    for lot in lots:
        origin_hint = guess_origin_name(lot.coffee_name) or guess_origin_name(lot.category or "")
        if not origin_hint:
            unresolved.append((lot.excel_row,
                               f"no origin match for {lot.coffee_name!r}"))
            continue
        origin_id = origin_by_name.get(origin_hint)
        if not origin_id:
            unresolved.append((lot.excel_row,
                               f"origin {origin_hint!r} not in MCR origins"))
            continue

        cs_id = find_coffee_source(cur, lot.coffee_name, origin_id, coffee_sources)
        if not cs_id:
            # Create a new coffee_source. coffee_name must be unique under (company, origin).
            cs_id = "csrc_" + uuid.uuid4().hex[:16]
            if not args.dry_run:
                cur.execute(
                    """
                    INSERT INTO coffee_source
                        (coffee_source_id, coffee_name, origin_id, company_id)
                    VALUES (%s, %s, %s, %s)
                    ON CONFLICT (company_id, origin_id, coffee_name)
                    DO UPDATE SET updated_at = now()
                    RETURNING coffee_source_id
                    """,
                    (cs_id, lot.coffee_name.strip(), origin_id, COMPANY_ID),
                )
                cs_id = cur.fetchone()["coffee_source_id"]
            coffee_sources.append({
                "coffee_source_id": cs_id,
                "coffee_name": lot.coffee_name.strip(),
                "origin_id": origin_id,
                "bag_size": None,
            })
            new_coffee_sources.append((cs_id, lot.coffee_name.strip()))

        # Resolve bag size as text matching bag_sizes table
        bag_size_int = int(round(lot.bag_size_lbs))
        bag_size_text = str(bag_size_int)
        if bag_size_text not in bag_size_ids:
            if not args.dry_run:
                cur.execute(
                    "INSERT INTO bag_sizes (bag_size_id, label, company_id) "
                    "VALUES (%s, %s, %s) ON CONFLICT (bag_size_id) DO NOTHING",
                    (bag_size_text, f"{bag_size_int} lb", COMPANY_ID),
                )
            bag_size_ids.add(bag_size_text)
            new_bag_sizes.append(bag_size_text)

        resolved.append({
            "lot": lot,
            "origin_id": origin_id,
            "origin_name": origin_hint,
            "coffee_source_id": cs_id,
            "bag_size_text": bag_size_text,
        })

    print(f"\nResolved {len(resolved)} lot rows. "
          f"Unresolved: {len(unresolved)}")
    for r, reason in unresolved:
        print(f"  UNRESOLVED row {r}: {reason}")

    if args.dry_run:
        print("\n--dry-run: not modifying the database.")
        conn.rollback()
        return 0

    # Delete existing baselines for MCR
    cur.execute(
        "DELETE FROM coffee_inventory_purchased "
        "WHERE company_id = %s AND entry_method = 'baseline'",
        (COMPANY_ID,),
    )
    print(f"\nDeleted {cur.rowcount} existing baseline rows for MCR.")

    # Insert new per-lot baselines
    inserted = 0
    skipped_dup = 0
    touched_origins: set[str] = set()

    for item in resolved:
        lot = item["lot"]
        origin_id = item["origin_id"]
        cs_id = item["coffee_source_id"]
        bag_size_text = item["bag_size_text"]
        lot_id = lot.lot_id

        # Idempotency: skip if (coffee_source_id, lot_id) baseline already exists
        cur.execute(
            "SELECT 1 FROM coffee_inventory_purchased "
            "WHERE company_id = %s AND entry_method = 'baseline' "
            "AND coffee_source_id = %s AND lot_id = %s LIMIT 1",
            (COMPANY_ID, cs_id, lot_id),
        )
        if cur.fetchone():
            skipped_dup += 1
            continue

        opid = str(uuid.uuid4())

        # The BEFORE trigger compute_coffee_purchase_amount will overwrite
        # `amount` to bags_ordered * coffee_source.bag_size. We want the row to
        # reflect actual on-hand lbs (bags*size + partial). Insert first with
        # bags_ordered, then UPDATE amount/remaining/bag_size in a second
        # statement that does NOT touch the trigger columns.
        cur.execute(
            """
            INSERT INTO coffee_inventory_purchased (
                origin_purchase_id, origin, facility_id, company_id,
                coffee_source_id, bags_ordered, cost_lb,
                lot_id, entry_method
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'baseline')
            """,
            (opid, origin_id, FACILITY_ID, COMPANY_ID,
             cs_id, lot.bag_count, lot.cost_lb, lot_id),
        )
        # Override amount and bag_size with actual lot totals (partial included)
        cur.execute(
            """
            UPDATE coffee_inventory_purchased
            SET amount = %s, remaining_lbs = %s, bag_size = %s
            WHERE origin_purchase_id = %s
            """,
            (lot.total_lbs, lot.total_lbs, bag_size_text, opid),
        )
        inserted += 1
        touched_origins.add(origin_id)

    print(f"Inserted {inserted} baseline rows. Skipped {skipped_dup} duplicates.")

    # Refresh origin stock totals
    for origin_id in touched_origins:
        cur.execute(
            "SELECT recalculate_origin_total_stock(%s, %s)",
            (origin_id, FACILITY_ID),
        )

    conn.commit()
    print(f"Refreshed total stock for {len(touched_origins)} origins.")

    # --- Audit summary -----------------------------------------------------
    print("\n=== AUDIT SUMMARY ===")
    print(f"Spreadsheet rows parsed: {len(lots)}")
    print(f"Spreadsheet rows skipped (zero on-hand / no cost): {len(skipped)}")
    print(f"Rows resolved to origin + source: {len(resolved)}")
    print(f"Rows unresolved: {len(unresolved)}")
    print(f"Rows inserted: {inserted}")
    print(f"Duplicates skipped (idempotent re-run): {skipped_dup}")

    print(f"\nNew coffee_source rows created: {len(new_coffee_sources)}")
    for cs_id, name in new_coffee_sources:
        print(f"  + {cs_id}  {name}")

    print(f"\nNew bag_sizes rows created: {len(new_bag_sizes)}")
    for b in new_bag_sizes:
        print(f"  + {b}")

    # Per-origin totals from the actually inserted rows
    print("\nPer-origin imported totals:")
    by_origin: dict[str, dict] = defaultdict(
        lambda: {"rows": 0, "bags": 0.0, "lbs": 0.0, "cost": 0.0}
    )
    for item in resolved:
        lot = item["lot"]
        d = by_origin[item["origin_name"]]
        d["rows"] += 1
        d["bags"] += lot.bag_count
        d["lbs"] += lot.total_lbs
        d["cost"] += lot.total_lbs * lot.cost_lb

    for name in sorted(by_origin):
        d = by_origin[name]
        print(f"  {name:20s} rows={d['rows']:3d} bags={d['bags']:7.2f} "
              f"lbs={d['lbs']:10.2f} cost=${d['cost']:11.2f}")

    # Final verification
    cur.execute(
        "SELECT count(*) AS c FROM coffee_inventory_purchased "
        "WHERE company_id = %s",
        (COMPANY_ID,),
    )
    total_rows = cur.fetchone()["c"]
    cur.execute(
        "SELECT count(*) AS c FROM coffee_inventory_purchased "
        "WHERE company_id = %s AND entry_method = 'baseline'",
        (COMPANY_ID,),
    )
    baseline_rows = cur.fetchone()["c"]
    cur.execute(
        "SELECT COALESCE(SUM(total_stock_lbs), 0) AS s "
        "FROM coffee_inventory WHERE company_id = %s",
        (COMPANY_ID,),
    )
    total_stock = float(cur.fetchone()["s"])

    sheet_total = sum(item["lot"].total_lbs for item in resolved)

    print(f"\ncoffee_inventory_purchased rows for MCR: {total_rows} "
          f"(baseline: {baseline_rows})")
    print(f"SUM(total_stock_lbs) for MCR coffee_inventory: {total_stock:.2f}")
    print(f"SUM(spreadsheet total_lbs) imported:           {sheet_total:.2f}")
    print(f"Delta: {total_stock - sheet_total:+.2f} lbs")

    if skipped:
        print(f"\nSkipped spreadsheet rows (first 20):")
        for r, reason in skipped[:20]:
            print(f"  row {r}: {reason}")

    cur.close()
    conn.close()
    return 0


if __name__ == "__main__":
    sys.exit(main())
