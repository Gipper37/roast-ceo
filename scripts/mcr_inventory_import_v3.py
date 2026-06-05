#!/usr/bin/env python3
"""
MCR April 2026 inventory import — v3 (schema-truth + locked rules).

Inputs:
  MCR/strata migration/April 2026 - Inventory Spreadsheet.xlsx
  MCR/strata migration/Wholesale Price List - March. 2026 Revised.pdf
Output:
  scripts/mcr_inventory_import_v3.sql

Target: company_id=9ShiyDAXhV, facility_id=5cc581b9-...

Depends on (NOT applied here):
  supabase/migrations/20260604000001_product_category.sql

Schema truths verified:
- consumable_type has ONLY two rows per company: 'product' (BOM) and
  'operational'. We resolve their UUIDs at runtime via a CTE-ish SELECT
  and reuse them. No new consumable_type rows are emitted.
- consumable_inventory.fallback_unit_cost ALREADY exists. No is_operational.
- coffee_inventory has latest_cost AND fallback_cost columns.
- coffee_source has NO cost column. Per-lot cost lives on
  coffee_inventory.fallback_cost (origin-aggregate mean of lot $/lb) and
  later on coffee_inventory_purchased rows from real shipment uploads.
  TODO: when invoice upload backfill lands, populate coffee_inventory_purchased
        per lot with real $/lb + bag_size + date_received.
- equipment_brand.category CHECK = espresso_machine|grinder|brewer|roaster|
  packaging|water_treatment|scale|other. We use 'other'.

Locked decisions implemented in v3:
- 11 product_category list:
    Hawaiian, Decaf (merged H+I), Blends, Custom Coffees, Flavor,
    International, Organic, Certified Organic (USDA), Sir Wilfreds,
    VIP, Equipment.
- "10%" prefix kept only when item contains "Blend" AND mentions a
  Hawaiian-entity token (Kau, Kona, Maui, Hawaiian, Hawaii, Hi, Lokelani,
  Hoala, Pacific, Moka, Mokka). Otherwise stripped.
- Consumable BOM/operational sheets:
    operational: SYRUPS & SAUCES, TEAS, PUMP POTS, ESPRESSO - CLEANERS,
                 BOXES, SUPPLIES
    product:     FLAVORING, BAGS, LABELS, SLEEVES
- Equipment items do NOT get a products row — they stay in equipment.
  The "Equipment" category_id is created for future product-list filtering
  but no product references it now.
- Everything else unchanged from v2.
"""

from __future__ import annotations

import hashlib
import re
import sys
from collections import defaultdict
from datetime import date
from pathlib import Path
from typing import Any, Iterable

import openpyxl

# ---------- constants ----------
COMPANY_ID = "9ShiyDAXhV"
FACILITY_ID = "5cc581b9-2803-42c2-98de-0ba16ae42f8e"
WHOLESALE_CHANNEL_ID = "6e6f4b92-8d17-4858-913a-b38b85b178a6"
RETAIL_CHANNEL_ID    = "87f69426-eb0f-4b67-a160-62c8be988323"
BASELINE_DATE = "2026-04-30"

SRC = Path("/Users/wanderingaloha/my-supabase-project/MCR/strata migration/April 2026 - Inventory Spreadsheet.xlsx")
OUT = Path("/Users/wanderingaloha/my-supabase-project/scripts/mcr_inventory_import_v3.sql")

# ---------- final 11 categories ----------
CATEGORIES = [
    "Hawaiian",
    "Decaf",
    "Blends",
    "Custom Coffees",
    "Flavor",
    "International",
    "Organic",
    "Certified Organic (USDA)",
    "Sir Wilfreds",
    "VIP",
    "Equipment",
]

# ---------- size config per oz sheet ----------
OZ_SIZES = {
    # sheet -> (size_name, weight_lbs, channels)
    "80oz": ("5lbs", 5.0,    ("wholesale",)),
    "32oz": ("2lbs", 2.0,    ("wholesale",)),
    "16oz": ("1lb",  1.0,    ("wholesale", "retail")),
    "8oz":  ("8oz",  0.5,    ("wholesale", "retail")),
    "7oz":  ("7oz",  0.4375, ("wholesale", "retail")),
    "2oz":  ("2oz",  0.125,  ("wholesale", "retail")),
}

CHANNEL_IDS = {
    "wholesale": WHOLESALE_CHANNEL_ID,
    "retail":    RETAIL_CHANNEL_ID,
}

# ---------- name normalization ----------
NORMALIZE_REPLACEMENTS = [
    (r"\bMac Nut\b", "Macnut"),
    (r"\bMacnut\b",  "Macnut"),
    (r"\bHazlenut\b","Hazelnut"),
    (r"\bHazulenut\b","Hazelnut"),
    (r"\bHazelnut\b","Hazelnut"),
    (r"\bLokalani\b","Lokelani"),
    (r"\bHawaiin\b", "Hawaiian"),
    (r"\bHawaiian Hazlenut\b","Hawaiian Hazelnut"),
    (r"\bMoglana\b", "Mogiana"),
    (r"\bGutamalan\b","Guatemalan"),
    (r"\bChoco\b",   "Chocolate"),
    (r"\bCarmel\b",  "Caramel"),
    (r"\bRasp\b",    "Raspberry"),
    (r"\bRasberry\b","Raspberry"),
    (r"\bPeaburry\b","Peaberry"),
    (r"\bPeabury\b", "Peaberry"),
    (r"\bMarcos\b",  "Marco's"),
    (r"\bMamas\b",   "Mama's"),
    (r"\bKau\b",     "Ka'u"),
    (r"Ka’u",        "Ka'u"),
]

def normalize(s: str) -> str:
    if not s:
        return s
    t = str(s).strip()
    t = re.sub(r"\s+", " ", t)
    for pat, rep in NORMALIZE_REPLACEMENTS:
        t = re.sub(pat, rep, t, flags=re.IGNORECASE)
    t = t.replace("marco's", "Marco's").replace("mama's", "Mama's")
    return t

# Hawaiian-entity tokens used by the corrected "10%" rule
HAWAIIAN_TOKENS = [
    r"\bka'u\b", r"\bkau\b", r"\bkona\b", r"\bmaui\b",
    r"\bhawaiian\b", r"\bhawaii\b", r"\bhi\b",
    r"\blokelani\b", r"\bhoala\b", r"\bpacific\b",
    r"\bmoka\b", r"\bmokka\b",
]
HAWAIIAN_RE = re.compile("|".join(HAWAIIAN_TOKENS), flags=re.IGNORECASE)

# Strip "10%" prefix unless BOTH (a) "Blend" present AND (b) Hawaiian token present.
def strip_10pct(name: str) -> str:
    if not name:
        return name
    has_blend = bool(re.search(r"\bblend\b", name, flags=re.IGNORECASE))
    has_hi    = bool(HAWAIIAN_RE.search(name))
    if has_blend and has_hi:
        return name
    return re.sub(r"^\s*10%\s*", "", name, flags=re.IGNORECASE).strip()

# Strip size prefix like "5lb", "1 Hr", "8oz", "1 hd", "2oz" — anything that
# repeats the size in the description. We want the "group name" portion.
SIZE_PREFIX_PATTERNS = [
    r"^\s*5\s*lb\s+",
    r"^\s*2\s*lb\s+",
    r"^\s*1\s*lb\s+",
    r"^\s*8\s*oz\s+",
    r"^\s*7\s*oz\s+",
    r"^\s*2\s*oz\s+",
    r"^\s*80\s*z\s+",
    r"^\s*1\s*Hr\s+",
    r"^\s*1\s*hd\s+",
    r"^\s*1\s*Fr\s+",
    r"^\s*1\s*Fd\s+",
    r"^\s*1\s*Org\s+",
    r"^\s*1\s*r\s+",
    r"^\s*1\s*d\s+",
    r"^\s*8\s*hr\s+",
    r"^\s*8\s*hd\s+",
    r"^\s*8\s*fr\s+",
    r"^\s*8\s*fd\s+",
    r"^\s*8\s*r\s+",
    r"^\s*8\s*d\s+",
]

def strip_size_prefix(name: str) -> str:
    if not name:
        return name
    t = name
    for p in SIZE_PREFIX_PATTERNS:
        t = re.sub(p, "", t, flags=re.IGNORECASE)
    return t.strip()

# ---------- group merge map ----------
GROUP_MERGES = {
    "Pacific Peaberry Blend": "Pacific Peaberry",
    "Pacific Peaberry":       "Pacific Peaberry",

    "Hawaii Blend":   "Hawaii Blend",
    "Hawaiian Blend": "Hawaii Blend",

    "BW Blend (Marco's Blend)": "Marco's Blend",
    "BW Blend (Marco's BLD)":   "Marco's Blend",
    "BW Blend (Marco's)":       "Marco's Blend",
    "BW Blend (Marcos)":        "Marco's Blend",
    "BW Blend":                 "Marco's Blend",
    "BW Blend Decaf":           "Marco's Blend",
    "Beans World Blend":        "Marco's Blend",
    "Beans World Blend Decaf":  "Marco's Blend",
    "Marco's Blend":            "Marco's Blend",
    "Marcos Blend":             "Marco's Blend",

    "Mama's Blend":             "Mama's Fish House Blend",
    "Mama's Fish House":        "Mama's Fish House Blend",
    "Mama's Fish House Blend":  "Mama's Fish House Blend",
}

def merge_group(name: str) -> str:
    return GROUP_MERGES.get(name, name)

# ---------- category routing ----------
# Inputs are the (sheet_category, normalized_group_name) — output is one of CATEGORIES.
def route_category(sheet_category: str | None, group_name: str) -> str:
    sc = (sheet_category or "").strip().upper()
    gn = group_name.lower()

    # Special-case items moved to "Custom Coffees":
    if "kraken" in gn or "maui resort rentals" in gn:
        return "Custom Coffees"
    if ("mama's fish house espresso" in gn or "red moka" in gn or "nobu" in gn
            or "cafe blend" in gn or "house blend" in gn
            or "pukalani superette" in gn or "french maui blend" in gn):
        if sc in ("DISCOUNTED PRICE", "DISCOUNTED PRICE TOTAL"):
            return "Custom Coffees"

    if sc in ("HAWAIIAN", "HAWAIIN", "HAWAIIAN "):
        if gn.startswith("hoala") or "lokelani" in gn:
            return "Blends"
        return "Hawaiian"
    # Decaf merge — both Hawaiian Decaf + International Decaf -> Decaf
    if sc in ("HAWAIIAN DECAF", "HAWAIIAN DECAF TOTAL",
              "INTERNATIONAL DECAF", "INTERNATIONAL DECAF "):
        return "Decaf"
    if sc in ("NEW BLENDS", "BLENDS", "MAUI BLEND", "MAUI BLD"):
        return "Blends"
    if sc in ("DISCOUNTED PRICE", "DISCOUNTED PRICE TOTAL"):
        return "Custom Coffees"
    # Flavor merge — Flavor + Flavor Decaf -> Flavor
    if sc in ("FLAVOR", "FLAVOR ", "FLAVOR DECAF", "FLAVORED DECAF"):
        return "Flavor"
    if sc in ("INTERNATIONAL", "INTERNATIONAL "):
        return "International"
    # Organic merge — International Organic -> Organic
    if sc in ("ORGANIC", "INTERNATIONAL ORGANIC", "INTERNATIONAL ORGANIC "):
        return "Organic"
    if sc in ("BOX - USDA", "BOX- USDA", "BOX-USDA"):
        return "Certified Organic (USDA)"
    if sc == "SW":
        return "Sir Wilfreds"
    if sc == "VIP":
        return "VIP"
    return "Hawaiian"

# ---------- wholesale price list (PDF) lookup ----------
PRICE_LIST = {
    "maui dark":          {"5lb": 185.75, "16oz": 37.15, "8oz": 18.58},
    "maui light":         {"5lb": 185.75, "16oz": 37.15, "8oz": 18.58},
    "maui medium":        {"8oz": 14.15},
    "nicky beans kona":   {"5lb": 292.50, "16oz": 58.50, "8oz": 29.25},
    "kona castaway":      {"5lb": 292.50, "16oz": 58.50, "8oz": 29.25},
    "kona peaberry":      {"5lb": 315.50, "16oz": 63.10, "8oz": 31.55},
    "maui moka goat":     {"5lb": 219.50, "16oz": 43.90, "8oz": 21.95},
    "maui moka":          {"5lb": 219.50, "16oz": 43.90, "8oz": 21.95},
    "maui red rooster":   {"5lb": 206.25, "16oz": 41.25, "8oz": 20.63},
    "maui yellow cat":    {"5lb": 206.25, "16oz": 41.25, "8oz": 20.63},
    "maui cat":           {"5lb": 206.25, "16oz": 41.25, "8oz": 20.63},
    "kona estate light":  {"5lb": 249.50, "16oz": 49.90, "8oz": 24.95},
    "kona estate dark":   {"5lb": 249.50, "16oz": 49.90, "8oz": 24.95},
    "ka'u":               {"5lb": 219.75, "16oz": 43.95, "8oz": 21.98},
    "hawaii blend":       {"5lb": 68.25, "16oz": 13.65, "8oz": 6.83},
    "hoala":              {"5lb": 73.75, "16oz": 14.75, "8oz": 7.38},
    "10% kona dark":      {"5lb": 77.50, "16oz": 15.50, "8oz": 7.75},
    "10% kona light":     {"5lb": 77.50, "16oz": 15.50, "8oz": 7.75},
    "lokelani":           {"5lb": 70.75, "16oz": 14.45, "8oz": 7.23},
    "10% maui blue bag":  {"8oz": 7.65},
    "10% maui":           {"5lb": 68.25, "16oz": 13.65, "8oz": 6.83},
    "maui blend":         {"5lb": 68.25, "16oz": 13.65, "8oz": 6.83},
    "no ka oi":           {"5lb": 73.75, "16oz": 14.75, "8oz": 7.38},
    "pacific":            {"5lb": 73.75, "16oz": 14.75, "8oz": 7.38},
    "pacific peaberry":   {"5lb": 73.75, "16oz": 14.75, "8oz": 7.38},
    "pacific blend":      {"5lb": 73.75, "16oz": 14.75, "8oz": 7.38},
    "espresso":           {"5lb": 63.25, "16oz": 12.65, "8oz": 6.33},
    "french roast":       {"5lb": 63.25, "16oz": 12.65, "8oz": 6.33},
    "anu anu":            {"5lb": 63.25},
    "anu anu cold brew":  {"5lb": 63.25},
    "organic dark":       {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic dark roast": {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic espresso":   {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic french":     {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic hapa":       {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic hapa blend": {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic light":      {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic light roast":{"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic medium":     {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic medium roast":{"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "organic blend":      {"5lb": 74.75, "16oz": 14.95, "8oz": 7.48},
    "baba's butterscotch":{"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "babas butterscotch": {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "chocolate macnut":   {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "coconut":            {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "hazelnut":           {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "hawaiian hazelnut":  {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "hula pie":           {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "macnut":             {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "vanilla macnut":     {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "vanilla":            {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "flavor":             {"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "chocolate raspberry":{"5lb": 60.75, "16oz": 12.15, "8oz": 6.08},
    "brazilian":          {"16oz": 12.15, "8oz": 6.08},
    "brazil mogiana":     {"16oz": 12.15, "8oz": 6.08},
    "colombian":          {"16oz": 12.15, "8oz": 6.08},
    "colombian supremo":  {"16oz": 12.15, "8oz": 6.08},
    "costa rican":        {"16oz": 12.15, "8oz": 6.08},
    "guatemalan":         {"16oz": 12.15, "8oz": 6.08},
    "honduras":           {"16oz": 12.15, "8oz": 6.08},
    "sumatra":            {"16oz": 12.15, "8oz": 6.08},
    "european":           {"16oz": 12.15, "8oz": 6.08},
    "fresh trades":       {"8oz": 6.24},
    "espresso decaf":     {"5lb": 62.00, "16oz": 12.40, "8oz": 6.20},
    "flavor decaf":       {"5lb": 64.25, "16oz": 12.85, "8oz": 6.43},
    "french roast decaf": {"5lb": 62.00, "16oz": 12.40, "8oz": 6.20},
    "hawaiian blend decaf":{"5lb": 67.00, "16oz": 13.40, "8oz": 6.70},
    "maui blend decaf":   {"5lb": 67.00, "16oz": 13.40, "8oz": 6.70},
    "club imua":          {"8oz": 6.24},
    "dawn patrol":        {"8oz": 6.24},
}

def lookup_price(group_name: str, sheet_name: str) -> float | None:
    key = group_name.lower().strip()
    key = re.sub(r"\s+", " ", key)
    sz_key = {"80oz":"5lb","32oz":"5lb","16oz":"16oz","8oz":"8oz","7oz":"8oz","2oz":"8oz"}[sheet_name]
    if key in PRICE_LIST and sz_key in PRICE_LIST[key]:
        return PRICE_LIST[key][sz_key]
    k2 = key.replace(" blend", "").strip()
    if k2 in PRICE_LIST and sz_key in PRICE_LIST[k2]:
        return PRICE_LIST[k2][sz_key]
    return None

# ---------- consumable sheets config ----------
# sheet -> (desc_col, qty_col, cost_col, bom_or_op, unit_label)
# bom_or_op: 'product' (BOM-eligible) or 'operational'
CONSUMABLE_SHEETS = {
    "FLAVORING":          (3, 8, 4, "product",     "lbs"),
    "BAGS":               (2, 6, 4, "product",     "case"),
    "LABELS":             (2, 5, 3, "product",     "roll"),
    "SYRUPS & SAUCES":    (3, 8, 4, "operational", "bottle"),
    "TEAS":               (3, 7, 4, "operational", "box"),
    "SLEEVES":            (2, 3, 6, "product",     "case"),
    "PUMP POTS":          (3, 7, 6, "operational", "each"),
    "ESPRESSO - CLEANERS":(3, 5, 4, "operational", "each"),
    "BOXES":              (1, 7, 8, "operational", "bundle"),
    "SUPPLIES":           (3, 6, 5, "operational", "case"),
}

# ---------- green coffee origin consolidation ----------
def consolidate_origin(lot_desc: str) -> str:
    s = lot_desc.lower()
    if "ka'u" in s or "kau " in s or s.strip() == "kau":
        return "Ka'u"
    if "kona" in s:
        return "Kona"
    if "maui" in s:
        return "Maui"
    if "hawaii no" in s or "hi3" in s or s.startswith("hawaii"):
        return "Hawaii"
    if "brazil" in s:
        return "Brazil"
    if "colombia" in s:
        return "Colombia"
    if "costa rica" in s:
        return "Costa Rica"
    if "guatemala" in s:
        return "Guatemala"
    if "el salvador" in s:
        return "El Salvador"
    if "honduras" in s:
        return "Honduras"
    if "nicaragu" in s:
        return "Nicaragua"
    if "mexico" in s or "veracruz" in s:
        return "Mexico"
    if "peru" in s:
        return "Peru"
    if "papa new guinea" in s or "png" in s:
        return "Papua New Guinea"
    if "sumatra" in s or "sulawesi" in s:
        return "Sumatra"
    if "yemen" in s:
        return "Yemen"
    if "timor" in s:
        return "Timor"
    return "Other"

# ---------- deterministic IDs ----------

def stable_id(prefix: str, *parts: str, length: int = 16) -> str:
    h = hashlib.sha1("||".join(parts).encode("utf-8")).hexdigest()[:length]
    return f"{prefix}_{h}"

def stable_uuid(*parts: str) -> str:
    h = hashlib.sha1("||".join(parts).encode("utf-8")).hexdigest()
    return f"{h[0:8]}-{h[8:12]}-4{h[13:16]}-8{h[17:20]}-{h[20:32]}"

# ---------- SQL helpers ----------

def sql_lit(v: Any) -> str:
    if v is None:
        return "NULL"
    if isinstance(v, bool):
        return "TRUE" if v else "FALSE"
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (v != v):
            return "NULL"
        return repr(v)
    s = str(v).replace("'", "''")
    return f"'{s}'"

def clean(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    return v

def to_num(v: Any) -> float | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v) if v == v else None
    s = str(v).strip().replace("$", "").replace(",", "")
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None

# ---------- load workbook ----------
print(f"Reading {SRC}", file=sys.stderr)
wb = openpyxl.load_workbook(SRC, data_only=True)

sql_chunks: list[str] = []
stats: dict[str, dict[str, int]] = {}

# trackers
created_categories: dict[str, str] = {}   # name -> id
created_groups:     dict[str, str] = {}
created_sizes:      dict[str, str] = {}
created_bag_sizes:  set[str] = set()
created_brands:     dict[str, str] = {}
created_models:     dict[tuple[str, str], str] = {}
created_products:   set[str] = set()
created_origins:    dict[str, str] = {}
created_sources:    set[str] = set()
consumable_counts:  dict[str, dict[str, int]] = defaultdict(lambda: {"product": 0, "operational": 0})

green_agg: dict[str, dict] = defaultdict(lambda: {
    "bags": 0.0, "costs": [], "bag_sizes": defaultdict(int), "lots": []
})

# ---------- ensure helpers ----------

def ensure_category(name: str) -> str:
    if name in created_categories:
        return created_categories[name]
    cid = stable_id("pcat", "mcr", name)
    created_categories[name] = cid
    sort = CATEGORIES.index(name) if name in CATEGORIES else 99
    sql_chunks.append(
        "INSERT INTO product_category (id, company_id, name, sort_order) "
        f"VALUES ({sql_lit(cid)}, {sql_lit(COMPANY_ID)}, {sql_lit(name)}, {sort}) "
        "ON CONFLICT (company_id, name) DO NOTHING;"
    )
    return cid

def ensure_size(name: str, weight: float) -> str:
    if name in created_sizes:
        return created_sizes[name]
    if name == "5lbs":
        sid = "34e787be"
        created_sizes[name] = sid
        return sid
    sid = stable_id("size", "mcr", name)
    created_sizes[name] = sid
    sql_chunks.append(
        "INSERT INTO size (size_id, size_name, weight, company_id) "
        f"VALUES ({sql_lit(sid)}, {sql_lit(name)}, {sql_lit(weight)}, {sql_lit(COMPANY_ID)}) "
        "ON CONFLICT (size_id) DO NOTHING;"
    )
    return sid

def ensure_product_group(name: str) -> str:
    key = name.strip()
    if key in created_groups:
        return created_groups[key]
    gid = stable_uuid("mcr_group", key)
    created_groups[key] = gid
    sql_chunks.append(
        "INSERT INTO product_groups (group_id, group_name, company_id, facility_id) "
        f"VALUES ({sql_lit(gid)}, {sql_lit(key)}, {sql_lit(COMPANY_ID)}, {sql_lit(FACILITY_ID)}) "
        "ON CONFLICT (group_id) DO NOTHING;"
    )
    return gid

def ensure_bag_size(bag_size_id: str) -> str:
    if bag_size_id in created_bag_sizes:
        return bag_size_id
    created_bag_sizes.add(bag_size_id)
    sql_chunks.append(
        "INSERT INTO bag_sizes (bag_size_id, label, company_id, facility_id) "
        f"VALUES ({sql_lit(bag_size_id)}, {sql_lit(bag_size_id)}, {sql_lit(COMPANY_ID)}, {sql_lit(FACILITY_ID)}) "
        "ON CONFLICT (bag_size_id) DO NOTHING;"
    )
    return bag_size_id

def ensure_brand(name: str, category: str = "other") -> str:
    key = name.strip()
    if key in created_brands:
        return created_brands[key]
    bid = stable_id("brand", "mcr", key.lower())
    created_brands[key] = bid
    sql_chunks.append(
        "INSERT INTO equipment_brand (equipment_brand_id, category, name, company_id) "
        f"VALUES ({sql_lit(bid)}, {sql_lit(category)}, {sql_lit(key)}, {sql_lit(COMPANY_ID)}) "
        "ON CONFLICT ON CONSTRAINT equipment_brand_scope_name_uq DO NOTHING;"
    )
    return bid

def ensure_model(brand_name: str, brand_id: str, model_name: str, category: str = "other") -> str:
    key = (brand_name.strip(), model_name.strip())
    if key in created_models:
        return created_models[key]
    mid = stable_id("model", "mcr", brand_name.lower(), model_name.lower())
    created_models[key] = mid
    sql_chunks.append(
        "INSERT INTO equipment_model (equipment_model_id, brand_id, category, model_name, company_id) "
        f"VALUES ({sql_lit(mid)}, {sql_lit(brand_id)}, {sql_lit(category)}, {sql_lit(model_name)}, {sql_lit(COMPANY_ID)}) "
        "ON CONFLICT ON CONSTRAINT equipment_model_scope_uq DO NOTHING;"
    )
    return mid

# ---------- GREEN COFFEE ----------

def process_green_coffee():
    sheet = wb["GREEN COFFEE"]
    parsed = 0
    sql_chunks.append("\n-- ===== GREEN COFFEE =====")
    for r in range(4, sheet.max_row + 1):
        row = [clean(sheet.cell(row=r, column=c).value) for c in range(1, sheet.max_column + 1)]
        item = row[0]
        if not isinstance(item, str):
            continue
        il = item.strip().lower()
        if "total" in il and (row[3] is None or row[4] is None) and (len(item.split()) <= 4):
            continue
        if il in ("hawaiian", "maui", "decaf", "international", "international organic", "international decaf"):
            continue
        supplier = row[3]
        cost_lb = to_num(row[4])
        bag_lbs = to_num(row[5])
        bags    = to_num(row[6]) or 0
        partial = to_num(row[7])
        if supplier is None and cost_lb is None and bag_lbs is None and not bags:
            continue
        parsed += 1
        origin_label = consolidate_origin(item)
        agg = green_agg[origin_label]
        agg["bags"] += bags
        if cost_lb is not None:
            agg["costs"].append(cost_lb)
        if bag_lbs is not None:
            agg["bag_sizes"][str(int(round(bag_lbs)))] += 1
        agg["lots"].append({
            "name": normalize(item),
            "supplier": supplier,
            "cost_lb": cost_lb,
            "bag_lbs": bag_lbs,
            "bags": bags,
            "partial": partial,
        })

    KNOWN_GLOBAL_BAGS = {"154","152","132","100","110","66"}
    for origin_label, agg in sorted(green_agg.items()):
        origin_id = stable_id("orig", "mcr", origin_label)
        created_origins[origin_label] = origin_id
        if agg["bag_sizes"]:
            bag_size_id = max(agg["bag_sizes"].items(), key=lambda x: x[1])[0]
            if bag_size_id not in KNOWN_GLOBAL_BAGS:
                ensure_bag_size(bag_size_id)
        else:
            bag_size_id = None
        if agg["costs"]:
            avg_cost = round(sum(agg["costs"]) / len(agg["costs"]), 4)
        else:
            avg_cost = None
        sql_chunks.append(
            "INSERT INTO coffee_inventory (origin_id, origin, company_id, facility_id, "
            "bag_size, inventory_count_bags, last_inventory, latest_cost, fallback_cost, is_active) "
            f"VALUES ({sql_lit(origin_id)}, {sql_lit(origin_label)}, {sql_lit(COMPANY_ID)}, {sql_lit(FACILITY_ID)}, "
            f"{sql_lit(bag_size_id)}, {sql_lit(agg['bags'])}, DATE {sql_lit(BASELINE_DATE)}, "
            f"{sql_lit(avg_cost)}, {sql_lit(avg_cost)}, TRUE) "
            "ON CONFLICT (origin_id) DO NOTHING;"
        )
        for lot in agg["lots"]:
            cs_id = stable_id("csrc", "mcr", origin_label, lot["name"], str(lot.get("supplier") or ""))
            if cs_id in created_sources:
                continue
            created_sources.add(cs_id)
            lot_bag_size = None
            if lot["bag_lbs"] is not None:
                lbs = str(int(round(lot["bag_lbs"])))
                if lbs not in KNOWN_GLOBAL_BAGS:
                    ensure_bag_size(lbs)
                lot_bag_size = lbs
            farm = lot["supplier"] or ""
            sql_chunks.append(
                "INSERT INTO coffee_source (coffee_source_id, coffee_name, origin_id, "
                "bag_size, company_id, farm, country_of_origin, is_active) "
                f"VALUES ({sql_lit(cs_id)}, {sql_lit(lot['name'])}, {sql_lit(origin_id)}, "
                f"{sql_lit(lot_bag_size)}, {sql_lit(COMPANY_ID)}, {sql_lit(farm)}, "
                f"{sql_lit(origin_label)}, TRUE) "
                "ON CONFLICT ON CONSTRAINT uq_coffee_source DO NOTHING;"
            )
    stats["GREEN COFFEE"] = {"parsed": parsed,
                             "origins": len(green_agg),
                             "sources": len(created_sources)}

# ---------- PRODUCTS ----------

def derive_group_from_desc(desc: str) -> str:
    g = strip_size_prefix(desc)
    g = normalize(g)
    g = strip_10pct(g)
    g = re.sub(r"\s+", " ", g).strip()
    parts = []
    for w in g.split():
        if w.startswith("10%"):
            parts.append(w)
        else:
            parts.append(w[:1].upper() + w[1:] if w else w)
    g = " ".join(parts)
    g = merge_group(g)
    return g

def process_oz_sheet(sheet_name: str):
    sheet = wb[sheet_name]
    size_name, weight_lbs, channels = OZ_SIZES[sheet_name]
    size_id = ensure_size(size_name, weight_lbs)
    cost_col = 6 if sheet_name == "80oz" else 4
    qty_col, desc_col = 5, 3
    parsed = inserted = 0
    current_category = None
    sql_chunks.append(f"\n-- ===== PRODUCTS: {sheet_name} ({size_name}) =====")
    for r in range(4, sheet.max_row + 1):
        row = [clean(sheet.cell(row=r, column=c).value) for c in range(1, sheet.max_column + 1)]
        if not any(v is not None for v in row):
            continue
        col1 = row[0]
        desc = row[desc_col - 1] if desc_col - 1 < len(row) else None
        if col1 is not None and desc is None:
            c1l = str(col1).strip().lower()
            if "total" in c1l or c1l.startswith("sum"):
                continue
            current_category = str(col1).strip()
            continue
        if not isinstance(desc, str):
            continue
        dl = desc.strip().lower()
        if "total" in dl:
            continue
        if sheet_name == "2oz" and desc.strip().upper() == "ITEM DESCRIPTION":
            continue
        if sheet_name == "8oz" and (current_category or "").strip().upper() == "VIP":
            if "mama" not in dl:
                continue
        parsed += 1
        cost = to_num(row[cost_col - 1]) if cost_col - 1 < len(row) else None
        qty  = to_num(row[qty_col  - 1]) if qty_col  - 1 < len(row) else None
        group_name = derive_group_from_desc(desc)
        if not group_name:
            continue
        cat_name = route_category(current_category, group_name)
        cat_id = ensure_category(cat_name)
        group_id = ensure_product_group(group_name)
        for ch_name in channels:
            ch_id = CHANNEL_IDS[ch_name]
            product_id = stable_id("prod", "mcr", sheet_name, ch_name, group_name)
            if product_id in created_products:
                continue
            created_products.add(product_id)
            price = lookup_price(group_name, sheet_name) if ch_name == "wholesale" else None
            sql_chunks.append(
                "INSERT INTO products (product_id, company_id, facility_id, group_id, category_id, "
                "size, channel, weight_lbs, price, is_active) "
                f"VALUES ({sql_lit(product_id)}, {sql_lit(COMPANY_ID)}, {sql_lit(FACILITY_ID)}, "
                f"{sql_lit(group_id)}, {sql_lit(cat_id)}, {sql_lit(size_id)}, {sql_lit(ch_id)}, "
                f"{sql_lit(weight_lbs)}, {sql_lit(price)}, TRUE) "
                "ON CONFLICT (product_id) DO NOTHING;"
            )
            inserted += 1
    stats[sheet_name] = {"parsed": parsed, "inserted": inserted}

# ---------- CONSUMABLES ----------

# Resolve the two existing consumable_type rows for MCR at runtime via psql vars.
# consumable_type is now a global lookup table (migration 20260605000001).
# Two stable IDs, no per-tenant resolution needed.
def emit_consumable_type_resolver():
    sql_chunks.append(
        "\n-- consumable_type IDs are global "
        "(see migration 20260605000001_consumable_type_global.sql).\n"
    )

def consumable_type_sql(bom_or_op: str) -> str:
    return (
        "'global_consumable_type_product'"
        if bom_or_op == "product"
        else "'global_consumable_type_operational'"
    )

def process_consumable_sheet(sheet_name: str):
    sheet = wb[sheet_name]
    desc_col, qty_col, cost_col, bom_or_op, _unit = CONSUMABLE_SHEETS[sheet_name]
    parsed = inserted = 0
    sql_chunks.append(f"\n-- ===== CONSUMABLE: {sheet_name} (type={bom_or_op}) =====")

    is_syrups = sheet_name == "SYRUPS & SAUCES"
    current_supplier = "Monin" if is_syrups else None

    for r in range(4, sheet.max_row + 1):
        row = [clean(sheet.cell(row=r, column=c).value) for c in range(1, sheet.max_column + 1)]
        if not any(v is not None for v in row):
            continue
        if is_syrups:
            col1 = row[0]
            if isinstance(col1, str) and col1.strip():
                c1u = col1.strip().upper()
                if "TOTAL" in c1u:
                    continue
                if c1u.startswith("MONIN"):
                    current_supplier = "Monin SF" if "SUGAR FREE" in c1u or "SF" in c1u else "Monin"
                    continue
                if c1u.startswith("GUITTARD"):
                    current_supplier = "Guittard"
                    continue
                if c1u.startswith("BRILLIANCE"):
                    current_supplier = "Brilliance"
                    continue
        desc = row[desc_col - 1] if desc_col - 1 < len(row) else None
        if not (isinstance(desc, str) and desc.strip()):
            if desc_col >= 2 and desc_col - 2 < len(row):
                alt = row[desc_col - 2]
                if isinstance(alt, str) and alt.strip():
                    desc = alt
        if not isinstance(desc, str):
            continue
        dl = desc.strip().lower()
        if "total" in dl or dl.startswith("sum "):
            continue
        cost = to_num(row[cost_col - 1]) if cost_col - 1 < len(row) else None
        qty  = to_num(row[qty_col  - 1]) if qty_col  - 1 < len(row) else None
        if cost is None and qty is None:
            continue
        parsed += 1
        item_name = normalize(desc)
        notes_supplier = current_supplier
        if is_syrups:
            if re.search(r"bril+iance", item_name, flags=re.IGNORECASE):
                notes_supplier = "Brilliance"
            elif notes_supplier == "Brilliance":
                notes_supplier = "Monin"
        full_name = item_name
        if is_syrups and notes_supplier and notes_supplier.lower() not in item_name.lower():
            full_name = f"{notes_supplier} {item_name}"
        cid = stable_id("cons", "mcr", sheet_name, full_name)
        baseline = qty if qty is not None else 0
        sql_chunks.append(
            "INSERT INTO consumable_inventory (consumable_inventory_id, consumable_inventory_item, "
            "company_id, facility_id, consumable_type, inventory_count, last_inventory_date, "
            "last_cost_unit, fallback_unit_cost, is_active) "
            f"VALUES ({sql_lit(cid)}, {sql_lit(full_name)}, {sql_lit(COMPANY_ID)}, {sql_lit(FACILITY_ID)}, "
            f"{consumable_type_sql(bom_or_op)}, {sql_lit(baseline)}, DATE {sql_lit(BASELINE_DATE)}, "
            f"{sql_lit(cost)}, {sql_lit(cost)}, TRUE) "
            "ON CONFLICT (consumable_inventory_id) DO NOTHING;"
        )
        inserted += 1
        consumable_counts[sheet_name][bom_or_op] += 1
    stats[sheet_name] = {"parsed": parsed, "inserted": inserted}

# ---------- EQUIPMENT ----------

def process_equipment():
    sheet = wb["EQUIPMENT"]
    parsed = inserted = 0
    current_brand: str | None = None
    sql_chunks.append("\n-- ===== EQUIPMENT =====")
    for r in range(4, sheet.max_row + 1):
        row = [clean(sheet.cell(row=r, column=c).value) for c in range(1, 6)]
        if not any(v is not None for v in row):
            continue
        make, model, qty, cost, _total = row
        if isinstance(make, str):
            ml = make.strip().lower()
            if "total" in ml:
                continue
            current_brand = make.strip()
            if model is None:
                continue
        if not isinstance(model, str):
            continue
        if model.strip().lower() in ("total:", "total"):
            continue
        if current_brand is None:
            continue
        parsed += 1
        brand_id = ensure_brand(current_brand, category="other")
        model_id = ensure_model(current_brand, brand_id, model, category="other")
        q = int(to_num(qty) or 1)
        unit_cost = to_num(cost)
        for i in range(max(q, 1)):
            eq_id = stable_id("eq", "mcr", current_brand.lower(), model.lower(), str(i))
            note = f"Imported from April 2026 inventory. Cost/unit: {unit_cost}" if unit_cost else "Imported from April 2026 inventory"
            sql_chunks.append(
                "INSERT INTO equipment (equipment_id, company_id, facility_id, brand_id, model_id, "
                "category, status, is_active, notes) "
                f"VALUES ({sql_lit(eq_id)}, {sql_lit(COMPANY_ID)}, {sql_lit(FACILITY_ID)}, "
                f"{sql_lit(brand_id)}, {sql_lit(model_id)}, 'other', 'active', TRUE, "
                f"{sql_lit(note)}) "
                "ON CONFLICT (equipment_id) DO NOTHING;"
            )
            inserted += 1
    stats["EQUIPMENT"] = {"parsed": parsed, "inserted": inserted}

# ---------- run ----------
sql_chunks.append("-- MCR April 2026 inventory import (v3 — schema-truth + locked rules)")
sql_chunks.append("-- Generated by scripts/mcr_inventory_import_v3.py")
sql_chunks.append(f"-- Target company_id={COMPANY_ID}, facility_id={FACILITY_ID}")
sql_chunks.append("-- REQUIRES migration 20260604000001_product_category.sql applied first.")
sql_chunks.append("-- Idempotent: every INSERT is ON CONFLICT DO NOTHING with stable IDs.")
sql_chunks.append("--")
sql_chunks.append("-- TODO: 2oz Maui Blend has both 80-ct ($156) and 40-ct ($78) packs in source.")
sql_chunks.append("--       Currently merged into ONE product. Revisit when pack-count modeling is added.")
sql_chunks.append("-- TODO: coffee_source has no per-lot cost column. Per-lot $/lb will land later via")
sql_chunks.append("--       coffee_inventory_purchased rows (real shipment invoices). For now, the")
sql_chunks.append("--       origin-level fallback_cost on coffee_inventory is the only $/lb stamp.")
sql_chunks.append("BEGIN;")

# 1. Seed categories first
sql_chunks.append("\n-- ===== PRODUCT CATEGORIES (11 final, incl. Equipment) =====")
for cat in CATEGORIES:
    ensure_category(cat)

# 2. Resolve existing consumable_type IDs (no inserts)
emit_consumable_type_resolver()

# 3. Green coffee (origins + sources)
process_green_coffee()

# 4. Products (oz sheets)
for s in OZ_SIZES:
    process_oz_sheet(s)

# 5. Consumables
for s in CONSUMABLE_SHEETS:
    process_consumable_sheet(s)

# 6. Equipment
process_equipment()

sql_chunks.append("\nCOMMIT;")

OUT.write_text("\n".join(sql_chunks) + "\n")

# ---------- summary ----------
print("\n=== MCR Inventory Import v3 Summary ===")
print(f"Output: {OUT}")
print()
print(f"{'Sheet':<22} {'parsed':>8} {'inserted':>10}")
print("-" * 44)
total_p = total_i = 0
for sh, st in stats.items():
    p = st.get("parsed", 0)
    i = st.get("inserted", st.get("origins", 0))
    print(f"{sh:<22} {p:>8} {i:>10}")
    total_p += p; total_i += i
print("-" * 44)
print(f"{'TOTAL':<22} {total_p:>8} {total_i:>10}")
print()
print(f"Reference rows emitted:")
print(f"  product_category : {len(created_categories)}")
print(f"  product_groups   : {len(created_groups)}")
print(f"  size             : {len(created_sizes)}")
print(f"  bag_sizes        : {len(created_bag_sizes)}")
print(f"  equipment_brand  : {len(created_brands)}")
print(f"  equipment_model  : {len(created_models)}")
print(f"  green origins    : {len(green_agg)}")
print(f"  coffee_source    : {len(created_sources)}")
print(f"  products         : {len(created_products)}")

op_ct = bom_ct = 0
for sh, counts in consumable_counts.items():
    bom_ct += counts["product"]
    op_ct  += counts["operational"]
print(f"  consumables (operational FK) : {op_ct}")
print(f"  consumables (product/BOM FK) : {bom_ct}")
